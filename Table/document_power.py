import os
import openpyxl as xl
from openpyxl.styles import Alignment, PatternFill, Border, Side



def make_data(values):
    data = f"""
            Company Name: {values['Company Name']}
            Address: {values['Address']}
            Trader: {values['Trader']}
            Trade Date: {values['Trade Date']}
            UOM Conversion: {values['UOM Conversion']},
            Delivery Date From: {values['Delivery Date From']},
            Delivery Date To: {values['Delivery Date To']},
            Volume: {values['Volume']},
            Frequency: {values['Frequency']}
            """
    return data



def create_powerTestSheet():

    # Assuming document_user.py is in the 'Table' directory
    current_directory = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(current_directory, 'excel_power.xlsx')
    # print(path)
    wb = xl.load_workbook(path)

    sheet_to_delete = 'Sheet1'
    if sheet_to_delete in wb.sheetnames:
        sheet = wb[sheet_to_delete]
        wb.remove(sheet)
        
    wb.create_sheet('Sheet1')
    sheet = wb['Sheet1']

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.row_dimensions[1].height = 30

    all_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    center_allignment = Alignment(horizontal='center', vertical='center')

    gray = PatternFill(start_color='A0A0A0', end_color='A0A0A0', fill_type='solid')

    id_cell = sheet.cell(1, 1)
    id_cell.value = "ID"
    id_cell.alignment = center_allignment
    id_cell.border = all_border
    id_cell.fill = gray
    
    test_data_cell = sheet.cell(1, 2)
    test_data_cell.value = "Form Inputs"
    test_data_cell.alignment = center_allignment
    test_data_cell.border = all_border
    test_data_cell.fill = gray

    validity_cell = sheet.cell(1, 3)
    validity_cell.value = "Validity"
    validity_cell.alignment = center_allignment
    validity_cell.border = all_border
    validity_cell.fill = gray

    status_cell = sheet.cell(1, 4)
    status_cell.value = "status"
    status_cell.alignment = center_allignment
    status_cell.border = all_border
    status_cell.fill = gray

    wb.save(path)



def add_powerTestSheet(values):

    # Assuming document_user.py is in the 'Table' directory
    current_directory = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(current_directory, 'excel_power.xlsx')
    wb = xl.load_workbook(path)
    sheet = wb['Sheet1']

    center_allignment = Alignment(horizontal='center', vertical='center')

    green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red = PatternFill(start_color='FF4651', end_color='FF4651', fill_type='solid')
    yellow = PatternFill(start_color='F0D700', end_color='F0D700', fill_type='solid')
    blue = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    near_white = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    all_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    id_cell = sheet.cell(sheet.max_row+1, 1)
    id_cell.value = values["id"]
    id_cell.alignment = center_allignment
    id_cell.border = all_border
    id_cell.fill = near_white

    form_data = make_data(values)
    test_data_cell = sheet.cell(sheet.max_row, 2)
    test_data_cell.value = form_data
    test_data_cell.alignment = xl.styles.Alignment(wrap_text=True)
    test_data_cell.border = all_border
    test_data_cell.fill = near_white

    validity_cell = sheet.cell(sheet.max_row, 3)
    validity_cell.value = values["validity"]
    validity_cell.alignment = center_allignment
    validity_cell.border = all_border
    if values['validity'] == 1:
        validity_cell.fill = yellow
    else:
        validity_cell.fill = blue

    status_cell = sheet.cell(sheet.max_row, 4)
    status_cell.value = values['status']
    status_cell.alignment = center_allignment
    status_cell.border = all_border
    if values['status'] == "pass":
        status_cell.fill = green
    else:
        status_cell.fill = red

    sheet.row_dimensions[sheet.max_row].height = 150

    wb.save(path) 


"""
values = {
      "id": "TC_Add_000",
      "Company Name": "KBS Energy",
      "Address": "",
      "Trader": "A Energy",
      "Trade Date": "01/16/2024",
      "UOM Conversion": "Kw",
      "Delivery Date From": "01/17/2024",
      "Delivery Date To": "01/18/2024",
      "Volume": "1",
      "Frequency": "Daily",
      "validity": 0,
      "status": "pass"
    }
"""

#create_powerTestSheet()
#add_powerTestSheet(values)
#add_powerTestSheet(values)
#add_powerTestSheet(values)
#add_powerTestSheet(values)